from pandas import read_excel
import pandas as pd
from shutil import copy2
import glob
import xlwings as xw
import re
import time
import tqdm
import openpyxl
import sys
import os


temp_file_to_wrote = {
    "report file- apps column name": "",
    "order excel - sheet with all apps under test": "",
    "pass_unpass": "",
    "error_type": "",
    "alert_type": "",
    "folder name": ""
}


class Compatibility:

    def __init__(self, file_name):
        self.file_name_ = file_name
        self.header_list = []
        self.__temp_count = 0

    @property
    def Temp_count(self):
        return self.__temp_count

    @Temp_count.setter
    def Temp_count(self, value):
        if self:
            self.__temp_count = value

    @Temp_count.deleter
    def Temp_count(self):
        self.__temp_count = 0

    def check_header(self, path2, sheet):
        """Checking header lp. in current sheet - tree"""
        print("Current data:", self.__temp_count)
        df = read_excel(path2, sheet_name=sheet, index_col=0, header=self.__temp_count)
        keys = df.head().keys()

        data_to_compare = temp_file_to_wrote["report file- apps column name"]
        if data_to_compare in keys:
            print("Keys match! LP. ", self.Temp_count)
            # print(f"Header in this sheet is on the {index_keys}")
            self.header_list.append(df)
            del self.Temp_count

        else:
            print("No matching keys in line: ", self.Temp_count)
            self.Temp_count += 1
            self.check_header(path2, sheet)

    @staticmethod
    def update_base_line_excel(path2, read_xlsx, read_app_base):
        # updating progress bar
        range_i = int(len(read_xlsx))
        pbar = tqdm.tqdm(total=range_i, desc='Start Copy Excel Sheets to new excel')
        StartTime = time.time()

        # copy Excel 1:1 from sheet 1 using excel mechanism
        try:
            for i in read_xlsx:
                if i in read_app_base:
                    wb = openpyxl.load_workbook(filename=i)
                    sheet = wb.sheetnames
                    print(f"\n\nAll sheet in {temp_file_to_wrote["order excel - sheet with all apps under test"]}", sheet)
                    if temp_file_to_wrote["order excel - sheet with all apps under test"] in sheet:
                        # print("\n\n\n\n\n\n\t", i, "\n\n\n\n\n\n\t")
                        index_temp = sheet.index(temp_file_to_wrote["order excel - sheet with all apps under test"])
                        # print("sheet.index(i)", index_temp)
                        wb1 = xw.Book(i)
                        wb2 = xw.Book(path2)

                        # +1 because starts from wx.Book form 1
                        ws1 = wb1.sheets(index_temp+1)
                        ws1.api.Copy(Before=wb2.sheets(1).api)
                        wb2.save()
                        wb2.app.quit()
                        pbar.update()

                else:
                    pbar.set_description("Processing %s\n" % i, refresh=None)
                    # print(path1)

                    wb1 = xw.Book(i)
                    wb2 = xw.Book(path2)

                    ws1 = wb1.sheets(1)
                    ws1.api.Copy(Before=wb2.sheets(1).api)
                    wb2.save()
                    wb2.app.quit()
                    pbar.update()
            pbar.close()

        except FileNotFoundError:
            print("There is no files in file- report")
            time.sleep(2)
            sys.exit()

        else:
            print('\n\n\tThat took {} seconds'.format(time.time() - StartTime))

    @staticmethod
    def resource_path(relative_path):
        """ Get absolute path to resource """
        try:
            # Creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
            print("Try", base_path)

        except Exception:
            base_path = os.path.abspath(".")
            print("Exception", base_path)

        return os.path.join(base_path, relative_path)

    @classmethod
    def print_file(cls, file_name):
        """Current file"""
        get_file_path = cls.resource_path(file_name)
        return get_file_path

    @staticmethod
    def get_excel_name(name):
        """Get name for the new report summary"""
        new_name = ""
        for i in name:
            if i == name[-1]:
                temp = (i[1:-1])
                try:
                    int(temp)

                except ValueError:
                    new_name += i
            else:
                new_name += i

        return new_name

    def mian(self):

        file_name = self.print_file(self.file_name_)
        print(file_name)

        read_xlsx = glob.glob('report\*.xlsx')
        read_app_base1 = glob.glob(f'*{temp_file_to_wrote["folder name"]}*.xlsx')
        read_app_base2 = glob.glob(f'*{temp_file_to_wrote["order excel - sheet with all apps under test"]}*.xlsx')
        # print(read_app_base1, read_app_base2)

        read_app_base = []
        read_app_base.extend(read_app_base1)
        read_app_base.extend(read_app_base2)
        print("\n\tRead app base: ", read_app_base)

        if len(read_xlsx) == 0:
            print("There is no reports Excel on 'report' direction"
                  "Copy there excel with results - for one project")
            time.sleep(5)
            sys.exit()

        if len(read_app_base) == 0:
            temp_name = temp_file_to_wrote["folder name"]
            print(f"There is no Excel order for all applications in the current direction\n"
                  f"It should be named eg. {temp_name}.xlsx; {temp_name}.xlsx")
            time.sleep(5)
            sys.exit()

        # print(read_app_base)
        name = re.findall(r"\[\w+]", read_xlsx[-1])
        time_format = time.strftime("_%Y_%m_%d_report")

        # Get Excel name for report file
        new_name = self.get_excel_name(name=name)

        path2 = new_name + time_format + ".xlsx"
        # Regular expression created name from square brake
        print(path2)

        read_xlsx.append(read_app_base[-1])
        print("read_xlsx", read_xlsx)

        print("\nBase excel copy...")
        try:
            copy2(file_name, path2)
        except FileNotFoundError:
            print("There is no Base.xlsx in folder")
            time.sleep(3)
            sys.exit()

        except PermissionError:
            print("\n\t! Close all excel file and start the program again")
            time.sleep(3)
            sys.exit()

        # UpdateBaseLineExcel
        self.update_base_line_excel(path2, read_xlsx, read_app_base)

        # create workbook
        wb = openpyxl.load_workbook(filename=path2)
        sheet = wb.sheetnames
        print("\n\nAll sheet", sheet)

        indexing_sheet = {}
        report_sheet = []
        report_sheet_app_list = None
        index_ = 0
        for i in sheet:
            if "reports" in i:
                report_sheet.append(i)

            elif "App List" in i:
                report_sheet_app_list = i

            # order for testing in main order excel file (all apps?) ????? >><<
            elif "Sheet1" in i:
                print("\n\n\n\n\n\n\t", i, "\n\n\n\n\n\n\t")
                report_sheet_app_list = i

            elif temp_file_to_wrote["order excel - sheet with all apps under test"] in i:
                print("\n\n\n\n\n\n\t", i, "\n\n\n\n\n\n\t")
                report_sheet_app_list = i

            index_ += 1

        print("\n\nReport sheet", report_sheet)

        # auto count header in report sheet
        for i in report_sheet:
            self.check_header(path2, i)

        data = self.header_list
        df_all_app = read_excel(path2, sheet_name=report_sheet_app_list, header=0)

        for key in df_all_app.keys():
            key_ = key.lower()

            if "apk name" in key_:
                keys = key
            elif "app package name" in key_:
                keys = key
            elif "package name" in key_:
                keys = key

        count_all_app = int(df_all_app[keys].count())
        print("\ncounted row:", count_all_app)

        result = pd.concat(data)
        coun_data = result.count()

        print("\nCount: ", coun_data)

        keys = []
        key_formatted = {"pass_unpass": None, "error_type": None, "alert_type": None}
        for key in result.keys():
            keys.append(key)

            if temp_file_to_wrote["pass_unpass"] in key:
                if "NA-NA" not in key:
                    print(key)
                    key_formatted["pass_unpass"] = key
            elif temp_file_to_wrote["error_type"] in key:
                if "NA-NA" not in key:
                    print(key)
                    key_formatted["error_type"] = key
            elif temp_file_to_wrote["alert_type"] in key:
                if "NA-NA" not in key:
                    print(key)
                    key_formatted["alert_type"] = key

        count_ = int(result[temp_file_to_wrote["report file- apps column name"]].count())

        key_format_pass_unpass = {}
        key_format_error_type = {}
        key_format_alert_type = {}

        for key, value in key_formatted.items():
            if key == "pass_unpass":
                count_lp = result[value].value_counts()
                # print("\n\n\tCount {}\n".format(key), count_lp)
                key_format_pass_unpass.update(count_lp)
                print("\nCount value key_format_pass_unpass\n", key_format_pass_unpass)

            elif key == "error_type":
                count_lp = result[value].value_counts()

                key_format_error_type.update(count_lp)
                print("\nCount value key_format_error_type\n", key_format_error_type)

            elif key == "alert_type":
                count_lp = result[value].value_counts()
                key_format_alert_type.update(count_lp)
                print("\nCount value key_format_alert_type\n", key_format_alert_type)

        pass_value = int(key_format_pass_unpass["pass"])
        pass_unpass = int(key_format_pass_unpass["unpass"])
        print("\n\npass_value", pass_value)
        print("pass_unpass", pass_unpass)

        # count
        unavailable = count_all_app - pass_value - pass_unpass
        print("unavailable", unavailable)
        # print("unavailable", unavailable)
        PassedRate = (pass_value / count_) * 100
        PassedRateTotal = (pass_value / count_all_app) * 100
        # print("PassedRate", PassedRate)
        PassedRate_ = round(PassedRate, 2)
        PassedRateTotal_ = round(PassedRateTotal, 2)
        print("\n\nPassedRate", PassedRate_, "\nPassedRateTotal_", PassedRateTotal_)

        error_type_txt = {'export_excel_crash_error': 0, 'export_excel_start_error': 0,
                          'export_excel_flash_out_error': 0, 'export_excel_install_error': 0,
                          'export_excel_white_black_error': 0, 'export_excel_notresp_error': 0}

        # dictionary to format txt all errors
        dict_format_key = {}
        dict_format_value = {}
        i = 0
        for key_, values_ in key_format_error_type.items():
            print("\n\n\tkey_", key_, "\nvalues_", values_)

            if 'export_excel_flash_out_error' in key_:
                error_type_txt['export_excel_flash_out_error'] += values_
                i += 1

            elif 'export_excel_white_black_error' in key_:
                error_type_txt['export_excel_white_black_error'] += values_
                i += 1

            elif 'export_excel_notresp_error' in key_:
                error_type_txt['export_excel_notresp_error'] += values_
                i += 1

            elif "export_excel_start_error" in key_:
                error_type_txt['export_excel_start_error'] += values_
                i += 1

            elif "export_excel_crash_error" in key_:
                error_type_txt['export_excel_crash_error'] += values_
                i += 1

            elif "export_excel_install_error" in key_:
                error_type_txt['export_excel_install_error'] += values_
                i += 1

            # maybe need be delete - kay can be too long
            else:
                if error_type_txt[key_] not in error_type_txt.keys():
                    error_type_txt[key_] = 1
                else:
                    error_type_txt[key_] += values_
                    i += 1

            dict_format_key[i] = key_
            dict_format_value[i] = values_

        index = len(dict_format_value)
        if index <= 20:
            print("\n\nAdding blank keys...")

            for i in range(index, 21):
                print("index without data", i)
                # last index + one!
                dict_format_key[i + 1] = ""
                dict_format_value[i + 1] = ""

        print("\n\tdict_format_value", dict_format_value, "\n\tdict_format_key", dict_format_key)
        # \t|{:^3}|   |{:^20}|".format(
        text = \
            f"""
We received list of {count_all_app} apps to perform tests, compatibility test has been done on {count_} 
apps.
Pass ratio on avaiable application: ({pass_value}/{count_})*100% = {PassedRate_}%. 


    >> {pass_unpass} cases failed:
            {error_type_txt["export_excel_start_error"]:>5}              {"Failed Start app":<}
            {error_type_txt["export_excel_crash_error"]:>5}              {"Not compatible with HMS":<}
            {error_type_txt["export_excel_flash_out_error"]:>5}             {"Unexpected exit":<}
            {error_type_txt[
        "export_excel_white_black_error"]:>5}              {"Black or white screen appeared":<}
            {error_type_txt["export_excel_notresp_error"]:>5}              {"App not response":<}
            {error_type_txt["export_excel_install_error"]:>5}              {"Install Failed":<}


                Detailed info from DevEco platform:
                Amount       Error
                    {dict_format_value[1]:>5}          {dict_format_key[1]:<}
                    {dict_format_value[2]:>5}          {dict_format_key[2]:<}
                    {dict_format_value[3]:>5}          {dict_format_key[3]:<}
                    {dict_format_value[4]:>5}          {dict_format_key[4]:<}
                    {dict_format_value[5]:>5}          {dict_format_key[5]:<}
                    {dict_format_value[6]:>5}          {dict_format_key[6]:<}
                    {dict_format_value[7]:>5}          {dict_format_key[7]:<}
                    {dict_format_value[8]:>5}          {dict_format_key[8]:<}
                    {dict_format_value[9]:>5}          {dict_format_key[9]:<}
                    {dict_format_value[10]:>5}          {dict_format_key[10]:<}
                    {dict_format_value[11]:>5}          {dict_format_key[11]:<}
                    {dict_format_value[12]:>5}          {dict_format_key[12]:<}
"""

        # wb = openpyxl.load_workbook(filename=path2)
        ws = wb['Test Report']
        ws['A7'] = count_all_app
        ws['B7'] = pass_value
        ws['C7'] = pass_unpass
        ws['D7'] = unavailable
        ws['E3'] = "{}%".format(PassedRateTotal_)
        ws['B5'] = time.strftime("(-%d.%m.%Y)")
        ws['B9'] = text

        if report_sheet_app_list == "Sheet1":
            ss_sheet = wb['Sheet1']
            ss_sheet.title = 'App List'
            print("renamed sheet....")

        elif report_sheet_app_list == temp_file_to_wrote["order excel - sheet with all apps under test"]:
            ss_sheet = wb[temp_file_to_wrote["order excel - sheet with all apps under test"]]
            ss_sheet.title = 'App List'
            print("renamed sheet....")
        else:
            pass

        wb.save(path2)


if __name__ == '__main__':
    # base file name to copy
    file_name = 'Base.xlsx'
    compatibility = Compatibility(file_name)
    compatibility.mian()
